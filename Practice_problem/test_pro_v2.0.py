import openpyxl


my_wb = openpyxl.Workbook()
my_sheet = my_wb.active

# my_wb.create_sheet(index = 2 , title = "Master_sheet")
# my_wb.save("D:\Git_Python\Practice_problem\Book1.xlsx")

my_path = "D:\Git_Python\Practice_problem\Book1.xlsx"
my_wb_obj = openpyxl.load_workbook(my_path)
my_sheet_obj = my_wb_obj.active
max_rows = my_sheet_obj.max_row
print(max_rows)
my_column = my_sheet_obj.max_column
print(my_column)

print("Enter the PS Number")
ps = int(input())
print("Enter the name : ")
name = input()
print("Enter the email Id")
email = input()
psFlag = 0
nameFlag = 0
emailFlag = 0
for i in range(1, max_rows + 1):
    cell_obj1 = my_sheet_obj.cell(row=i, column=1)
    if ps == cell_obj1.value:
        print(cell_obj1.value)
        psFlag = 1
        cell_obj1 = my_sheet_obj.cell(row=i, column=2)
        if cell_obj1.value == name:
            print(cell_obj1.value)
            nameFlag = 1
            cell_obj1 = my_sheet_obj.cell(row=i, column=3)
            if cell_obj1.value == email:
                print(cell_obj1.value)
                emailFlag = 1

if emailFlag == 1:
    print("Match found")


# Creating second sheet for inserting data
my_wb.create_sheet(index=0, title="Master")
my_wb.save("D:\Git_Python\Practice_problem\Book2.xlsx")

# Creating the data frame for it

temp =  my_sheet_obj.cell(row=1, column=1)
print(temp)
print(temp.value)
my_path1 = "D:\Git_Python\Practice_problem\Book2.xlsx"
my_wb_obj1 = openpyxl.load_workbook(my_path1)
mySheetObj1 = my_wb_obj1.active
c1 = mySheetObj1.cell(row = 1, column = 1)
c1.value = temp.value
my_wb.save("D:\Git_Python\Practice_problem\Book2.xlsx")







