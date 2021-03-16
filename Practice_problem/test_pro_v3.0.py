import openpyxl

path1 = "D:\Git_Python\Practice_problem\sample.xlsx"
book = openpyxl.load_workbook(path1)
sheet = book.active
print("Enter your PS Number :")
ps = int(input())
print("Enter your Name : ")
name = input()
print("Enter your Email : ")
email = (input())
maxRows = sheet.max_row
print(maxRows)
for i in range(1, maxRows+1):
    a1 = sheet.cell(row=i, column=1)
    if a1.value == ps:
        a2 = sheet.cell(row=i, column=2)
        if a2.value == name:
            a3 = sheet.cell(row=i, column=3)
            if a3.value == email:
                flag = 1
                print(a1.value, a2.value, a3.value)


