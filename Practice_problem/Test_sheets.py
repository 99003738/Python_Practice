# import openpyxl

# book = openpyxl.load_workbook('sheets.xlsx')
#
# print(book.get_sheet_names())
#
# active_sheet = book.active
# print(type(active_sheet))
#
# sheet = book.get_sheet_by_name("March")
# print(sheet.title)


# book = openpyxl.load_workbook('sheets.xlsx')
#
# book.create_sheet("April")
#
# print(book.sheetnames)
#
# sheet1 = book.get_sheet_by_name("January")
# book.remove_sheet(sheet1)
#
# print(book.sheetnames)
#
# book.create_sheet("January", 0)
# print(book.sheetnames)
#
# book.save('sheets2.xlsx')


import openpyxl


# opening the source excel file
filename = "D:/Git_Python/Practice_problem/sheets.xlsx"
wb1 = openpyxl.load_workbook(filename)
ws1 = wb1.worksheets[0]
mr = ws1.max_row
print(mr)
mc = ws1.max_column
print(mc)

# opening the destination excel file
filename1 = "D:/Git_Python/Practice_problem/Book1.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active

# calculate total number of rows and
# columns in source excel file

# copying the cell values from source
# excel file to destination excel file
for i in range(1, mr + 1):
    for j in range(1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row=i, column=j)
        print(c.value)
        # writing the read value to destination excel file
        ws2.cell(row=i, column=j).value = c.value

    # saving the destination excel file
wb2.save(str(filename1))
