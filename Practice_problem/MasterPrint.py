import openpyxl as xl
import UserInput
import DataFilter


def writing_mastersheet():
    pathMasterWorkbook = UserInput.outputPath[0]
    masterbook = xl.load_workbook(pathMasterWorkbook)
    lstsheet = masterbook.sheetnames
    length = len(lstsheet)
    currMassheet = masterbook.worksheets[0]
    print("\n")
    print("1. OverWrite ")
    print("2.New Sheet")
    choice = int(input("Enter Your choice : "))
    if choice == 1:
        currMassheet.delete_rows(2, currMassheet.max_row - 1)
        masterbook.save(UserInput.outputPath[0])
    elif choice == 2:
        ws1 = masterbook.create_sheet("Mysheet")
        masterbook.save(UserInput.outputPath[0])



def sending_data_master(pathvariable, searchitem = []):

    # creating workbook from where data to be searched
    workbook1 = xl.load_workbook(pathvariable)
    workbooksheet = workbook1.sheetnames
    lengthworkbook1 = len(workbooksheet)

    # opening the destination worksheet
    # pathMasterWorkbook = "C:\\Users\\Shivam\\Desktop\\MasterBook.xlsx"
    pathMasterWorkbook = UserInput.outputPath[0]
    masterbook = xl.load_workbook(pathMasterWorkbook)
    no_of_sheet = len(masterbook.sheetnames)
    # workSheetMaster = masterbook.active
    if no_of_sheet == 1:
        currMassheet = masterbook.worksheets[0]
    else:
        currMassheet = masterbook.worksheets[(no_of_sheet-1)]
    index = 0

    print(searchitem)
    for i in range(0, lengthworkbook1):
        sheets = workbook1.worksheets[i]
        # workSheetMaster = masterbook.active
        # calulating max row and max column

        wbMaxRow = sheets.max_row
        wbMaxColumn = sheets.max_column

        # taking length of the searchitem
        lengthSearch = len(searchitem)
        colnum = 1

        # finding the length of the master sheet
        mastermaxrow = currMassheet.max_row
        mastermaxcol = currMassheet.max_column

        rownum = mastermaxrow+2
        mastercol= 1
        for colnum in range(4, wbMaxColumn + 1):
            currMassheet.cell(row=rownum, column=mastercol).value = sheets.cell(row=1, column=colnum).value
            mastercol = mastercol+1
        mastercol= 1
        for i in range(2, wbMaxRow+1):
            data1 = searchitem[index]
            for j in range(1, 4):
                valuecheck = sheets.cell(row= i , column= j)
                value1 = str(valuecheck.value)
                # print(data1)
                # print(value1)
                if value1 == str(data1):
                    rownum = rownum + 1
                    mastercol =1
                    # print(data1)
                    # print(value1)
                    for k in range(4, wbMaxColumn+1):
                        cellprintpos = sheets.cell(row= i, column=k)
                        # print(cellprintpos)
                        cellprintval = cellprintpos.value
                        # print(cellprintval)
                        currMassheet.cell(row= rownum, column= mastercol).value =  sheets.cell(row=i, column= k).value
                        mastercol = mastercol+1
                    # index = index + 1
    masterbook.save(str(pathMasterWorkbook))


UserInput.user_selection()
UserInput.user_choice_selection()
UserInput.outputResult_path()
writing_mastersheet()
length0fpathlist = len(UserInput.pathList)
for d in range(0, length0fpathlist):
    path = UserInput.pathList[d]
    check = DataFilter.validating_input(path)
    if check == 1:
        sending_data_master(path, UserInput.inputList)
