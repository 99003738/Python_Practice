
'''
Importing the UserInput.py file
'''
import UserInput
'''
Importing the openpyxl library.
'''
from openpyxl import load_workbook
'''
Function name : sheettraversing()
Input :  sheet of a particular work book and search key from the global database.
Output : return 1 if input matched found in the sheet

"""
'''
def sheettraversing(sheet):
    # print("you are in sheettraversing")
    validFlag = 0
    maxrows = sheet.max_row   # counting the max row in the sheet
    maxcolumn = sheet.max_column
    # print(maxrows)
    # print(maxcolumn)
    loi_list = len(UserInput.inputList)   # counting length of the inputList
   # for listdata in range(0, loi_list):  rem
    for i in range(2, maxrows+1):                                   # i is rows and j is column
        for j in range(1,4):
             data_at_cell1 = sheet.cell(row= i, column= j)

             if str(UserInput.inputList[0]) == str(data_at_cell1.value):
                data_at_cell1 = sheet.cell(row= i, column = 1)  # loading all the data of matched search key
                data_at_cell2 = sheet.cell(row= i, column= 2)
                data_at_cell3 = sheet.cell(row= i, column= 3)
                v1 = str(data_at_cell1.value)
                v2 = str(data_at_cell2.value)
                v3 = str(data_at_cell3.value)
                v = [v1, v2, v3]
                # print(v)
                # print(loi_list)
                add =0
                for check in range(0, 3):
                    for inner in range(0, loi_list):
                        if v[check] == str(UserInput.inputList[inner]):
                            v[check] = 1
                            add = add + int(v[check])
                if add == loi_list:
                    print("Input matched found")
                    return 1
                    validFlag = 1

                    break
                break
        if validFlag ==1:
            break

'''
Function name : validating_input
Input :  taking path from the inputpath list entred by the user 
Output : if data found in any sheet then printing message 

"""
'''
def validating_input(path):
    # print("you are in validating input sect")
    # lenofinputpath = len(UserInput.pathList)

    # if UserInput.Dict["RF_ManualPath"] ==1 and UserInput.Dict["RF_ManualFilterInput"] ==1:
    #     print(lenofinputpath)
    # else:
    #     print("Exporting path error")
    # for i in range(0, lenofinputpath): # ? check the value here weather iterating correct or not
    count = 1
    # path = UserInput.pathList[i]
    workbook = load_workbook(path)
    sheetlist = workbook.sheetnames
    # print(sheetlist)
    totalsheet = len(sheetlist)
    # print(totalsheet)
    for sheet in range(0, totalsheet):
        print('you are in sheet', count)
        check = sheettraversing(workbook[sheetlist[sheet]])
        if check == 1:
            print("data found in sheet", count)
            return 1
            break
        else:
            count = count + 1
            print(" Sorry No record found ")


# UserInput.user_selection()
# UserInput.user_choice_selection()
# print(UserInput.inputList)
# validating_input(UserInput.pathList[0])