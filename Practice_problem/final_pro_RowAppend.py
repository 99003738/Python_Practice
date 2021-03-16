from openpyxl import load_workbook

wb = load_workbook('sheets.xlsx')
sheet1 = wb.get_sheet_by_name('sheet1')
sheet2 = wb.get_sheet_by_name('sheet2')
sheet3 = wb.get_sheet_by_name('sheet3')
sheet4 = wb.get_sheet_by_name('sheet4')
sheet5 = wb.get_sheet_by_name('sheet5')
Master = wb.get_sheet_by_name('Master')
maxRowsSheet1 = sheet1.max_row
maxColumnSheet1 = sheet1.max_column

maxRowsSheet2 = sheet2.max_row
maxColumnSheet2 = sheet2.max_column

maxRowsSheet3 = sheet3.max_row
maxColumnSheet3 = sheet3.max_column

maxRowsSheet4 = sheet4.max_row
maxColumnSheet4 = sheet4.max_column

maxRowsSheet5 = sheet5.max_row
maxColumnSheet5 = sheet5.max_column

print("Enter your PS Number :")
ps = int(input())
print("Enter your Name : ")
name = input()
print("Enter your Email : ")
email = (input())

flag = 0

for i in range(1, maxRowsSheet1 + 1):
    a1 = sheet1.cell(row=i, column=1)
    if a1.value == ps:
        a2 = sheet1.cell(row=i, column=2)
        if a2.value == name:
            a3 = sheet1.cell(row=i, column=3)
            if a3.value == email:
                flag = 1
                print(a1.value, a2.value, a3.value)
k = 2  # for rows
col = 1  # for column in master

if flag == 1:
    Master.cell(row=1, column=1).value = 'PS_Number'
    Master.cell(row=1, column=2).value = ps
    Master.cell(row=1, column=3).value = 'Name'
    Master.cell(row=1, column=4).value = name
    Master.cell(row=1, column=5).value = 'Email'
    Master.cell(row=1, column=6).value = email

    for r in range(4, maxColumnSheet1 + 1):
        Master.cell(row=2, column=col).value = sheet1.cell(row=1, column=r).value
        col = col + 1
    col = 1
    for i in range(1, maxRowsSheet1 + 1):
        a1 = sheet1.cell(row=i, column=1)
        a2 = sheet1.cell(row=i, column=2)
        a3 = sheet1.cell(row=i, column=3)
        if (a1.value == ps and a3.value == email) and a2.value == name:
            k = k + 1
            for j in range(4, maxColumnSheet1 + 1):
                Master.cell(row=k, column=col).value = sheet1.cell(row=i, column=j).value
                col = col + 1
            col = 1
    # this is for Sheet2
    maxRowMaster = Master.max_row
    col = 1
    maxRowMaster = maxRowMaster + 2
    for r in range(4, maxColumnSheet2 + 1):
        Master.cell(row=maxRowMaster, column=col).value = sheet2.cell(row=1, column=r).value
        col = col + 1

    for i in range(1, maxRowsSheet2 + 1):
        a1 = sheet2.cell(row=i, column=1)
        a2 = sheet2.cell(row=i, column=2)
        a3 = sheet2.cell(row=i, column=3)
        if (a1.value == ps and a3.value == email) and a2.value == name:
            col = 1
            maxRowMaster = maxRowMaster + 1
            for j in range(4, maxColumnSheet2 + 1):
                Master.cell(row=maxRowMaster, column=col).value = sheet2.cell(row=i, column=j).value
                col = col + 1
    # this is for Sheet3
    maxRowMaster = Master.max_row
    col = 1
    maxRowMaster = maxRowMaster + 2
    for r in range(4, maxColumnSheet3 + 1):
        Master.cell(row=maxRowMaster, column=col).value = sheet3.cell(row=1, column=r).value
        col = col + 1

    for i in range(1, maxRowsSheet3 + 1):
        a1 = sheet3.cell(row=i, column=1)
        a2 = sheet3.cell(row=i, column=2)
        a3 = sheet3.cell(row=i, column=3)
        if (a1.value == ps and a3.value == email) and a2.value == name:
            col = 1
            maxRowMaster = maxRowMaster + 1
            for j in range(4, maxColumnSheet3 + 1):
                Master.cell(row=maxRowMaster, column=col).value = sheet3.cell(row=i, column=j).value
                col = col + 1
    # this is for Sheet 4
    maxRowMaster = Master.max_row
    col = 1
    maxRowMaster = maxRowMaster + 2
    for r in range(4, maxColumnSheet4 + 1):
        Master.cell(row=maxRowMaster, column=col).value = sheet4.cell(row=1, column=r).value
        col = col + 1
    for i in range(1, maxRowsSheet4 + 1):
        a1 = sheet4.cell(row=i, column=1)
        a2 = sheet4.cell(row=i, column=2)
        a3 = sheet4.cell(row=i, column=3)
        if (a1.value == ps and a3.value == email) and a2.value == name:
            col = 1
            maxRowMaster = maxRowMaster + 1
            for j in range(4, maxColumnSheet4 + 1):
                Master.cell(row=maxRowMaster, column=col).value = sheet4.cell(row=i, column=j).value
                col = col + 1

    # this is for Sheet 5
    maxRowMaster = Master.max_row
    col = 1
    maxRowMaster = maxRowMaster + 2
    for r in range(4, maxColumnSheet5 + 1):
        Master.cell(row=maxRowMaster, column=col).value = sheet5.cell(row=1, column=r).value
        col = col + 1

    for i in range(1, maxRowsSheet5 + 1):
        a1 = sheet5.cell(row=i, column=1)
        a2 = sheet5.cell(row=i, column=2)
        a3 = sheet5.cell(row=i, column=3)
        if (a1.value == ps and a3.value == email) and a2.value == name:
            col = 1
            maxRowMaster = maxRowMaster + 1
            for j in range(4, maxColumnSheet5 + 1):
                Master.cell(row=maxRowMaster, column=col).value = sheet5.cell(row=i, column=j).value
                col = col + 1
    wb.save('sheets.xlsx')
else:
    print("N0 Record Found")

# coding to copy whole data
# for i in range(1, maxRows+1):
#     for j in range(1, sheet1.max_column+1):
#         sheet2.cell(row=i , column= j).value = sheet1.cell(row=i, column=j).value
# wb.save('sheets.xlsx')
