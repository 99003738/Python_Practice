import openpyxl







my_wb = openpyxl.Workbook()
my_sheet = my_wb.active
my_sheet_title = my_sheet.title
print("My sheet title: " + my_sheet_title)
my_sheet_title = "My New Sheet"
print("Sheet name is : " + my_sheet_title)

c1 = my_sheet.cell(row = 1, column = 1)
c1.value = "Aadrika"
c2 = my_sheet.cell(row= 1 , column = 2)
c2.value = "Adwaita"
c3 = my_sheet['A2']
c3.value = "Satyajit"
# B2 = column = 2 & row = 2.
c4 = my_sheet['B2']
c4.value = "Bivas"
my_wb.save("D:\Git_Python\Practice_problem\Book1.xlsx")

# Creating new sheet in the same sheet

my_wb.create_sheet(index = 1 , title = "New sheet")
my_wb.save("D:\Git_Python\Practice_problem\Book1.xlsx")

#display total number of rows

my_path = "D:\Git_Python\Practice_problem\Book1.xlsx"
my_wb_obj = openpyxl.load_workbook(my_path)
my_sheet_obj = my_wb_obj.active
print(my_sheet_obj.max_row)


#display a particular cell value

my_cell_obj = my_sheet_obj.cell(row = 1, column = 1)
print(my_cell_obj.value)

#display total number of column
print(my_sheet_obj.max_column)

# display all columns name

my_max_column = my_sheet_obj.max_column
for i in range(1,my_max_column + 1):
    my_cell_obj = my_sheet_obj.cell(row = 1, column = i)
    print(my_cell_obj.value)

# display first column value

my_row = my_sheet_obj.max_row
for i in range(1, my_row+1):
    cell_obj = my_sheet_obj.cell(row = i,column = 1)
    print(cell_obj.value)

#print an paricular row value

for i in range(1, my_max_column + 1):
    print("I am here")
    cell_obj = my_sheet_obj.cell(row = 2 , column = i)
    print(cell_obj.value, end = " ")