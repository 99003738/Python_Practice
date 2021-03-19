import openpyxl as xl
import UserInput
import DataFilter
from openpyxl import load_workbook

def search_by_input():
    length0fpathlist = len(UserInput.pathList)
    for d in range(0, length0fpathlist):
        path = UserInput.pathList[d]
        workbook1 = load_workbook(path)
        workbooksheet = workbook1.sheetnames
        totalsheets = len(workbooksheet)
        lenofinput = len(UserInput.inputList)
        storeprevdata = []
        for sheet in range(0, totalsheets):
            sheets = workbook1[workbooksheet[sheet]]
            for rows in range(2, sheets.max_row + 1):
                for col in range(1, 4):
                    prevdata = sheets.cell(row=rows, column=col).value

                    storeprevdata.append(prevdata)
                lenoflist = len(storeprevdata)
                add = 0
                for i in range(0, lenofinput):
                    for j in range(0, lenoflist):
                        if str(UserInput.inputList[i]) == str(storeprevdata[j]):
                            add = add + 1
                if add == lenofinput:
                    print(storeprevdata)
                storeprevdata = []





def searchdata():
    length0fpathlist = len(UserInput.pathList)
    for d in range(0, length0fpathlist):
        path = UserInput.pathList[d]
        workbook1 = xl.load_workbook(path)
        workbooksheet = workbook1.sheetnames
        totalsheets = len(workbooksheet)
        storeprevdata =[]
        for sheet in range(0, totalsheets):
            sheets = workbook1[workbooksheet[sheet]]
            storenextdata = []
            for rows in range(2, sheets.max_row+1):
                for col in range(1, 4):
                    prevdata = sheets.cell(row= rows, column= col).value
                    storenextdata.append(prevdata)
                    storeprevdata.append(prevdata)
                lenoflist = len(storeprevdata)
                print(lenoflist)
                lenofinput = len(storenextdata)
                add = 0
                i = 0
                for j in range(0, lenoflist):
                    if str(storenextdata[i]) == str(storeprevdata[j]):
                        if i<3:
                            i = i+1
                            add = add + 1
            if add == lenofinput:
                print(storenextdata)
                storenextdata= []
                # print(storenextdata)
                # print(storeprevdata)



UserInput.user_selection()
UserInput.user_choice_selection()
if UserInput.Dict["Flag_showData"] == 1:
    search_by_input()
elif UserInput.Dict["Flag_showData"] == 0:
    searchdata()

