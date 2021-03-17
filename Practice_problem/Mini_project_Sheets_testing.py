from openpyxl import load_workbook


######  This is code to take out the all the sheet name from the given workbook
print("reading directly the path")
path = "C:\\Users\\Shivam\\PycharmProjects\\Python_Practice\\Practice_problem\\sheets.xlsx"
workbook = load_workbook(path)
number_of_sheet = workbook.sheetnames
print(number_of_sheet)
#######################################################

# trying to get the multiple path in a list simultanously
print("multiple line path")
lines = []
while True:                                    # directly copy all the path in terminal with new line and hit enter and enter
    line = input()                             # the directory file will loaded into the list form and can be extract by list index
    if line:
        lines.append(line)
    else:
        break
print(lines)
print(len(lines))
text = '\n'.join(lines)
print(lines[0])
print(lines[1])
print(text)
######################################################################

# reading the path from text file
print("reading the path from file")
pathList = []
f = open("demo.txt", "r")
context = f.read()
CoList = context.split("\n")

print(CoList)
# code is working fine for reading through the file

###############################################################################

# NOW TAKING OUT THE PATH AND PRINTING THE SHEETS.
# issue is in the format of the path which taking the \\ between them.

pathVariable = CoList[0]
myWorkbook = load_workbook(pathVariable)
number_of_sheet = myWorkbook.sheetnames
print(number_of_sheet)                       # Tested OK
#####################################################################################